"""
Canon bootstrap document parsers.

Each parser returns a list of ParsedChunk objects extracted from the source document.
A chunk is a named section with raw text content and optional metadata.

Supported formats:
  - .md / .txt   → md_parser
  - .pdf         → pdf_parser  (requires pymupdf)
  - .xlsx / .xls → xlsx_parser (requires openpyxl)
  - .docx        → docx_parser (requires python-docx)
  - .html / .htm → html_parser (requires beautifulsoup4)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class ParsedChunk:
    source_file: str
    section: str         # heading or sheet name
    text: str            # raw extracted text
    page: int | None = None
    metadata: dict = field(default_factory=dict)


def parse_file(path: Path) -> list[ParsedChunk]:
    """Dispatch to the appropriate parser based on file extension."""
    suffix = path.suffix.lower()
    if suffix in (".md", ".txt"):
        return parse_md(path)
    elif suffix == ".pdf":
        return parse_pdf(path)
    elif suffix in (".xlsx", ".xls"):
        return parse_xlsx(path)
    elif suffix == ".docx":
        return parse_docx(path)
    elif suffix in (".html", ".htm"):
        return parse_html(path)
    else:
        logger.warning("Unsupported file type: %s — skipping", path.name)
        return []


def parse_directory(directory: Path) -> list[ParsedChunk]:
    """Parse all supported documents in a directory."""
    chunks: list[ParsedChunk] = []
    for p in sorted(directory.iterdir()):
        if p.is_file() and p.name != "README.md":
            try:
                file_chunks = parse_file(p)
                chunks.extend(file_chunks)
                logger.info("Parsed %s: %d chunks", p.name, len(file_chunks))
            except Exception as e:
                logger.warning("Failed to parse %s: %s", p.name, e)
    return chunks


# ── Markdown / plain text ──────────────────────────────────────────────────────

def parse_md(path: Path) -> list[ParsedChunk]:
    text = path.read_text(encoding="utf-8", errors="replace")
    chunks: list[ParsedChunk] = []
    current_section = "intro"
    current_lines: list[str] = []

    for line in text.splitlines():
        if line.startswith("#"):
            if current_lines:
                chunks.append(ParsedChunk(
                    source_file=path.name,
                    section=current_section,
                    text="\n".join(current_lines).strip(),
                ))
                current_lines = []
            current_section = line.lstrip("#").strip()
        else:
            current_lines.append(line)

    if current_lines:
        chunks.append(ParsedChunk(
            source_file=path.name,
            section=current_section,
            text="\n".join(current_lines).strip(),
        ))

    return [c for c in chunks if c.text]


# ── PDF ───────────────────────────────────────────────────────────────────────

def parse_pdf(path: Path) -> list[ParsedChunk]:
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning("PyMuPDF not installed; skipping PDF %s. Install with: uv add pymupdf", path.name)
        return []

    chunks: list[ParsedChunk] = []
    doc = fitz.open(str(path))
    for page_num, page in enumerate(doc, start=1):
        text = page.get_text("text").strip()
        if text:
            chunks.append(ParsedChunk(
                source_file=path.name,
                section=f"page_{page_num}",
                text=text,
                page=page_num,
            ))
    doc.close()
    return chunks


# ── Excel ─────────────────────────────────────────────────────────────────────

def parse_xlsx(path: Path) -> list[ParsedChunk]:
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed; skipping XLSX %s. Install with: uv add openpyxl", path.name)
        return []

    chunks: list[ParsedChunk] = []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers: list[str] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip fully empty rows
            if all(cell is None for cell in row):
                continue
            cells = [str(cell) if cell is not None else "" for cell in row]
            if row_idx == 0:
                headers = cells
            else:
                if headers:
                    rows.append(dict(zip(headers, cells)))
                else:
                    rows.append({"_row": "\t".join(cells)})

        if rows:
            # Convert to text — join as readable lines
            lines = []
            for r in rows[:200]:  # cap at 200 rows per sheet
                line = " | ".join(f"{k}: {v}" for k, v in r.items() if v and v != "None")
                if line:
                    lines.append(line)
            if lines:
                chunks.append(ParsedChunk(
                    source_file=path.name,
                    section=sheet_name,
                    text="\n".join(lines),
                    metadata={"headers": headers},
                ))

    wb.close()
    return chunks


# ── Word (DOCX) ───────────────────────────────────────────────────────────────

def parse_docx(path: Path) -> list[ParsedChunk]:
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx not installed; skipping DOCX %s. Install with: uv add python-docx", path.name)
        return []

    doc = Document(str(path))
    chunks: list[ParsedChunk] = []
    current_section = "intro"
    current_lines: list[str] = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        if para.style.name.startswith("Heading"):
            if current_lines:
                chunks.append(ParsedChunk(
                    source_file=path.name,
                    section=current_section,
                    text="\n".join(current_lines).strip(),
                ))
                current_lines = []
            current_section = text
        else:
            current_lines.append(text)

    if current_lines:
        chunks.append(ParsedChunk(
            source_file=path.name,
            section=current_section,
            text="\n".join(current_lines).strip(),
        ))

    return [c for c in chunks if c.text]


# ── HTML (Confluence exports) ─────────────────────────────────────────────────

def parse_html(path: Path) -> list[ParsedChunk]:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        logger.warning("beautifulsoup4 not installed; skipping HTML %s. Install with: uv add beautifulsoup4", path.name)
        return []

    html = path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    # Remove script/style noise
    for tag in soup(["script", "style", "nav", "header", "footer"]):
        tag.decompose()

    chunks: list[ParsedChunk] = []
    current_section = soup.title.string.strip() if soup.title else "content"
    current_lines: list[str] = []

    for tag in soup.find_all(["h1", "h2", "h3", "p", "li", "td", "th"]):
        text = tag.get_text(" ", strip=True)
        if not text:
            continue
        if tag.name in ("h1", "h2", "h3"):
            if current_lines:
                chunks.append(ParsedChunk(
                    source_file=path.name,
                    section=current_section,
                    text="\n".join(current_lines).strip(),
                ))
                current_lines = []
            current_section = text
        else:
            current_lines.append(text)

    if current_lines:
        chunks.append(ParsedChunk(
            source_file=path.name,
            section=current_section,
            text="\n".join(current_lines).strip(),
        ))

    return [c for c in chunks if c.text]
